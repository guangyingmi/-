#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Sheet 拆分工具（值+样式版）
- 自动搜索当前文件夹内的Excel文件，将每个Excel文件中的所有sheet拆分成独立的Excel文件
- 默认：去除公式，仅保留原表格“显示出来的值”，且尽量保持样式/布局一致
- 可选：--keep-formulas 可保留公式（恢复原行为）

依赖：openpyxl, tqdm
"""

import os
import sys
import datetime as _dt
import traceback
import argparse
from pathlib import Path
from typing import List
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm


def log(msg: str):
    """日志输出"""
    ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def sanitize_filename(name: str) -> str:
    """清理文件名中的非法字符"""
    import re
    # 替换Windows文件名中的非法字符
    return re.sub(r'[\\/:*?"<>|]', "_", str(name))


def copy_value_and_style(val_cell, style_cell, dst_cell):
    """
    把“值”和“样式”分别从不同源单元格拷到目标单元格：
    - 值：来自 data_only=True 的工作簿（即公式的缓存值）
    - 样式：来自 data_only=False 的工作簿
    """
    # 值（如果缓存为 None，就按 None 拷；不引入公式）
    dst_cell.value = val_cell.value

    # 样式
    if style_cell.has_style:
        dst_cell.font = copy(style_cell.font)
        dst_cell.border = copy(style_cell.border)
        dst_cell.fill = copy(style_cell.fill)
        dst_cell.number_format = style_cell.number_format
        dst_cell.protection = copy(style_cell.protection)
        dst_cell.alignment = copy(style_cell.alignment)


def copy_worksheet_values_and_styles(src_ws_vals, src_ws_styles, dst_ws):
    """复制工作表（仅值）+ 样式/布局"""
    # 复制所有单元格（值+样式）
    formula_count = 0
    missing_cached = 0

    for row in src_ws_styles.iter_rows():
        for style_cell in row:
            r, c = style_cell.row, style_cell.column
            val_cell = src_ws_vals.cell(row=r, column=c)
            if (val_cell.value is not None) or style_cell.has_style:
                dst_cell = dst_ws.cell(row=r, column=c)
                copy_value_and_style(val_cell, style_cell, dst_cell)

            # 统计无缓存值的公式（用于友好提示）
            try:
                is_formula = (
                    getattr(style_cell, "data_type", None) == "f" or
                    (isinstance(style_cell.value, str) and style_cell.value.startswith("="))
                )
                if is_formula:
                    formula_count += 1
                    if val_cell.value is None:
                        missing_cached += 1
            except Exception:
                pass

    # 复制列宽
    for col_letter, dimension in src_ws_styles.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dimension.width

    # 复制行高
    for row_num, dimension in src_ws_styles.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dimension.height

    # 复制合并单元格
    for merged_range in src_ws_styles.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # 冻结窗格
    if src_ws_styles.freeze_panes:
        dst_ws.freeze_panes = src_ws_styles.freeze_panes

    # 自动筛选
    if src_ws_styles.auto_filter:
        dst_ws.auto_filter.ref = src_ws_styles.auto_filter.ref

    # 打印设置
    try:
        dst_ws.page_setup = copy(src_ws_styles.page_setup)
        dst_ws.page_margins = copy(src_ws_styles.page_margins)
    except Exception:
        pass

    # 工作表保护
    try:
        if src_ws_styles.protection.sheet:
            dst_ws.protection = copy(src_ws_styles.protection)
    except Exception:
        pass

    return formula_count, missing_cached


def find_excel_files(directory: str) -> List[str]:
    """查找目录下的所有Excel文件"""
    excel_files = []
    directory_path = Path(directory)

    for file_path in directory_path.iterdir():
        if file_path.is_file():
            # openpyxl 不支持 .xls（老格式），这里仍然显示，但处理时会 try/except 跳过
            if file_path.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
                # 排除临时文件和输出文件夹
                if not file_path.name.startswith('~$') and '单个Excel合集' not in file_path.name:
                    excel_files.append(str(file_path))

    return excel_files


def split_excel_sheets(excel_file: str, output_dir: str, keep_formulas: bool) -> int:
    """拆分Excel文件中的所有sheet为独立文件"""
    try:
        log(f"正在处理文件: {os.path.basename(excel_file)}")

        # 过滤掉 .xls（openpyxl 不支持）
        if excel_file.lower().endswith(".xls"):
            log(f"  > 跳过 .xls 老格式：{os.path.basename(excel_file)}（建议另存为 .xlsx 再试）")
            return 0

        # 加载两个工作簿：
        # 1) 值工作簿（data_only=True）：用于拿到公式的缓存值
        # 2) 样式工作簿（data_only=False）：用于复制样式与布局
        if keep_formulas:
            # 保留公式的老行为：仅打开样式工作簿，并按原来的 copy 逻辑
            wb_styles = load_workbook(excel_file, data_only=False)
            wb_vals = None
        else:
            wb_vals = load_workbook(excel_file, data_only=True)
            wb_styles = load_workbook(excel_file, data_only=False)

        sheet_count = 0
        sheet_names = wb_styles.sheetnames
        log(f"发现 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")

        for sheet_name in tqdm(sheet_names, desc=f"拆分 {os.path.basename(excel_file)}"):
            try:
                src_ws_styles = wb_styles[sheet_name]
                if keep_formulas:
                    # 旧逻辑：保留公式（只拷贝一个源）
                    new_wb = Workbook()
                    new_wb.remove(new_wb.active)
                    new_ws = new_wb.create_sheet(title=sheet_name)

                    # 直接逐格复制（值+样式，值可能是公式）
                    for row in src_ws_styles.iter_rows():
                        for cell in row:
                            if cell.value is not None or cell.has_style:
                                dst_cell = new_ws.cell(row=cell.row, column=cell.column)
                                # 值：原值（可能是公式）
                                dst_cell.value = cell.value
                                # 样式
                                if cell.has_style:
                                    dst_cell.font = copy(cell.font)
                                    dst_cell.border = copy(cell.border)
                                    dst_cell.fill = copy(cell.fill)
                                    dst_cell.number_format = cell.number_format
                                    dst_cell.protection = copy(cell.protection)
                                    dst_cell.alignment = copy(cell.alignment)

                    # 复制其余布局项
                    for col_letter, dimension in src_ws_styles.column_dimensions.items():
                        new_ws.column_dimensions[col_letter].width = dimension.width
                    for row_num, dimension in src_ws_styles.row_dimensions.items():
                        new_ws.row_dimensions[row_num].height = dimension.height
                    for merged_range in src_ws_styles.merged_cells.ranges:
                        new_ws.merge_cells(str(merged_range))
                    if src_ws_styles.freeze_panes:
                        new_ws.freeze_panes = src_ws_styles.freeze_panes
                    if src_ws_styles.auto_filter:
                        new_ws.auto_filter.ref = src_ws_styles.auto_filter.ref
                    try:
                        new_ws.page_setup = copy(src_ws_styles.page_setup)
                        new_ws.page_margins = copy(src_ws_styles.page_margins)
                    except Exception:
                        pass
                    try:
                        if src_ws_styles.protection.sheet:
                            new_ws.protection = copy(src_ws_styles.protection)
                    except Exception:
                        pass

                else:
                    # 新逻辑：去除公式，仅保留值 + 样式
                    src_ws_vals = wb_vals[sheet_name]
                    new_wb = Workbook()
                    new_wb.remove(new_wb.active)
                    new_ws = new_wb.create_sheet(title=sheet_name)

                    formula_count, missing_cached = copy_worksheet_values_and_styles(
                        src_ws_vals, src_ws_styles, new_ws
                    )
                    if formula_count > 0 and missing_cached > 0:
                        log(f"  > 提示：工作表 '{sheet_name}' 中有 {missing_cached}/{formula_count} 个公式无缓存值（可能从未在Excel/WPS中计算过），导出处将为空。")

                # 生成输出文件名
                safe_sheet_name = sanitize_filename(sheet_name)
                output_file = os.path.join(output_dir, f"{safe_sheet_name}.xlsx")
                counter = 1
                original_output_file = output_file
                while os.path.exists(output_file):
                    name_without_ext = os.path.splitext(original_output_file)[0]
                    output_file = f"{name_without_ext}_{counter}.xlsx"
                    counter += 1

                # 保存文件
                new_wb.save(output_file)
                new_wb.close()

                sheet_count += 1
                log(f"已保存: {os.path.basename(output_file)}")

            except Exception as e:
                log(f"处理工作表 '{sheet_name}' 时出错: {str(e)}")
                continue

        # 关闭
        wb_styles.close()
        if not keep_formulas and wb_vals is not None:
            wb_vals.close()

        return sheet_count

    except Exception as e:
        log(f"处理文件 '{excel_file}' 时出错: {str(e)}")
        return 0


def main():
    """主函数"""
    try:
        parser = argparse.ArgumentParser(description="Excel Sheet 拆分（去除公式，仅保留值）")
        parser.add_argument("--keep-formulas", action="store_true", help="保留公式（默认去除公式，仅保留值）")
        args = parser.parse_args()

        log("Excel Sheet 拆分工具启动")
        log(f"当前工作目录: {os.getcwd()}")
        if args.keep_formulas:
            log("模式：保留公式（与旧版一致）")
        else:
            log("模式：去除公式，仅保留值（默认）")

        # 查找Excel文件
        excel_files = find_excel_files(os.getcwd())
        if not excel_files:
            log("未找到任何Excel文件")
            return

        log(f"找到 {len(excel_files)} 个Excel文件:")
        for file in excel_files:
            log(f"  - {os.path.basename(file)}")

        # 创建输出目录
        output_dir = os.path.join(os.getcwd(), "单个Excel合集")
        os.makedirs(output_dir, exist_ok=True)
        log(f"输出目录: {output_dir}")

        # 处理每个Excel文件
        total_sheets = 0
        successful_files = 0

        for excel_file in excel_files:
            try:
                sheet_count = split_excel_sheets(excel_file, output_dir, keep_formulas=args.keep_formulas)
                if sheet_count > 0:
                    total_sheets += sheet_count
                    successful_files += 1
                    log(f"文件 '{os.path.basename(excel_file)}' 处理完成，拆分了 {sheet_count} 个工作表")
                else:
                    log(f"文件 '{os.path.basename(excel_file)}' 处理失败或无可拆分工作表")
            except Exception as e:
                log(f"处理文件 '{os.path.basename(excel_file)}' 时发生错误: {str(e)}")
                continue

        # 输出总结
        log("\n=== 处理完成 ===")
        log(f"成功处理文件数: {successful_files}/{len(excel_files)}")
        log(f"总共拆分工作表数: {total_sheets}")
        log(f"输出目录: {output_dir}")

        if total_sheets > 0:
            log("所有工作表已成功拆分为独立的Excel文件！")
        else:
            log("没有成功拆分任何工作表")

    except Exception as e:
        error_msg = f"程序执行出错: {str(e)}\n{traceback.format_exc()}"
        log(error_msg)
        raise


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n用户中断程序")
    except Exception as e:
        log(f"程序异常退出: {str(e)}")
        sys.exit(1)
