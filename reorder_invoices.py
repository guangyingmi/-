#!/usr/bin/env python3
"""
reorder_invoices.py
-----------------------------------------------------------
双击后自动完成：
1. 扫描当前文件夹及子目录全部 PDF
2. 遍历 *.xls / *.xlsx：
      • prefix = Excel 文件名前缀（应收账款/预付账款/…）
      • 同一公司首次出现时，把该公司所有 PDF 一次性并入
      • 缺票 → Excel 行整行标红
      • 统计：PDF 总数、Excel 重复公司、PDF 多票公司、PDF 未匹配公司
3. 输出文件（每张 Excel 各自一组）：
      <prefix>-排序后合并PDF文件.pdf
      <prefix>-标记处理后Excel文件.xlsx
      <prefix>-未匹配PDF清单.txt
   以及全局：
      未使用到的PDF合并后文件.pdf
-----------------------------------------------------------
打包单文件（在 macOS / Windows 各自环境里）：
  pyinstaller --onefile reorder_invoices.py
"""

# ---------- 让打包后的可执行在自身目录运行 ----------
import os, sys, pathlib
if getattr(sys, 'frozen', False):
    os.chdir(pathlib.Path(sys.executable).parent)

from pathlib import Path
import re, unicodedata, collections
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm
from pypdf import PdfReader, PdfWriter

ROOT = Path.cwd()  # 当前运行目录

# ---------- 通用处理 ----------
def norm(txt: str | None) -> str:
    if txt is None:
        return ""
    # 全角→半角，去空格
    txt = unicodedata.normalize("NFKC", str(txt))
    txt = re.sub(r"\s+", "", txt)
    # 删除常见括号/标点
    txt = re.sub(r"[()（）【】\[\]「」『』“”\"'、，．·•\-—]", "", txt)
    return txt.upper()

def company_from_subject(subject: str) -> str:
    """Excel 科目名称 >>> 公司名"""
    return norm(subject.split("_", 1)[-1])

def company_from_filename(stem: str) -> str:
    """票文件名 >>> 公司名"""
    parts = stem.split("_")
    if len(parts) == 2:
        return norm(parts[1])
    for p in reversed(parts):
        if not re.fullmatch(r"\d{14}", p):  # 非 14 位时间戳
            return norm(p)
    return norm(stem)

def is_pdf_file(p: Path) -> bool:
    if p.suffix.lower() == ".pdf":
        return True
    try:
        with p.open("rb") as f:
            return f.read(5) == b"%PDF-"
    except Exception:
        return False

# ---------- 扫描全部 PDF ----------
def build_pdf_map(root: Path) -> dict[str, list[Path]]:
    pdf_map: dict[str, list[Path]] = {}
    for f in root.rglob("*"):
        if f.is_file() and is_pdf_file(f):
            pdf_map.setdefault(company_from_filename(f.stem), []).append(f)
    print(f"[INFO] 扫描到 PDF 总数: {sum(len(v) for v in pdf_map.values())}")
    return pdf_map

# ---------- 处理单张 Excel ----------
def process_excel(excel_path: Path, pdf_map: dict[str, list[Path]]):
    prefix = excel_path.stem.split("_", 1)[0]           # 应收账款 / 预付账款
    out_pdf   = ROOT / f"{prefix}-排序后合并PDF文件.pdf"
    out_excel = ROOT / f"{prefix}-标记处理后Excel文件.xlsx"
    out_txt   = ROOT / f"{prefix}-未匹配PDF清单.txt"

    print(f"\n=== 处理 {excel_path.name} ===")
    df = pd.read_excel(excel_path)
    if "科目名称" not in df.columns:
        print("❌ 缺少『科目名称』列，跳过")
        return

    # 统计重复
    counter = collections.Counter(company_from_subject(c) for c in df["科目名称"])
    dup_excel = {k: v for k, v in counter.items() if v > 1}
    if dup_excel:
        print("[INFO] Excel 重复公司:", ", ".join(f"{k}×{v}" for k, v in dup_excel.items()))

    writer = PdfWriter()
    wb = load_workbook(excel_path)
    ws = wb.active
    red = PatternFill("solid", fgColor="FFFF0000")
    seen, missing_rows, used = set(), [], set()

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="合并 PDF"):
        comp = company_from_subject(row["科目名称"])
        if comp in seen:
            continue
        seen.add(comp)
        bucket = pdf_map.get(comp, [])
        if bucket:
            used.add(comp)
            for pdf in bucket:
                for pg in PdfReader(pdf).pages:
                    writer.add_page(pg)
        else:
            r = idx + 2
            missing_rows.append(r)
            for cell in ws[r]:
                cell.fill = red

    # 主合并 PDF
    if writer.pages:
        with out_pdf.open("wb") as f:
            writer.write(f)
        print(f"[OK] 输出合并 PDF: {out_pdf.name} ({len(writer.pages)} 页)")
    wb.save(out_excel)
    print(f"[OK] 输出标红 Excel: {out_excel.name}")

    # 未匹配统计
    unmatched = {k: v for k, v in pdf_map.items() if k not in used and k not in counter}
    if unmatched:
        with out_txt.open("w", encoding="utf-8") as f:
            total = sum(len(v) for v in unmatched.values())
            f.write(f"未匹配 PDF 总计 {total} 张\n\n")
            for comp, files in sorted(unmatched.items(), key=lambda kv: len(kv[1]), reverse=True):
                f.write(f"{comp} ({len(files)} 张)\n")
                for p in files:
                    f.write(f"  {p}\n")
                f.write("\n")
        print(f"[INFO] 未匹配 PDF 清单: {out_txt.name}")

    if missing_rows:
        print("[WARN] 缺票行:", missing_rows)

    return unmatched  # 返回未匹配 dict 供后续全局合并

# ---------- 主入口 ----------
if __name__ == "__main__":
    pdf_map_full = build_pdf_map(ROOT)
    excel_paths = list(ROOT.glob("*.xls*"))
    if not excel_paths:
        print("⚠️ 当前目录未找到任何 Excel")
        sys.exit(0)

    # 用于收集所有 Excel 未匹配票
    unmatched_global: dict[str, list[Path]] = {}

    for x in excel_paths:
        unmatched = process_excel(x, pdf_map_full)
        for k, v in unmatched.items():
            unmatched_global.setdefault(k, []).extend(v)

    # 合并所有未匹配 PDF
    if unmatched_global:
        out_unmatched_pdf = ROOT / "未使用到的PDF合并后文件.pdf"
        writer_u = PdfWriter()
        for comp, files in sorted(unmatched_global.items()):
            for pdf in files:
                for pg in PdfReader(pdf).pages:
                    writer_u.add_page(pg)
        with out_unmatched_pdf.open("wb") as f:
            writer_u.write(f)
        print(f"\n[OK] 全部未匹配 PDF 已合并: {out_unmatched_pdf.name}")

    print("\n✅ 全部 Excel 处理完成")
