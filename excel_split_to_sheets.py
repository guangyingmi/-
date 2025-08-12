#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按人名把汇总表拆成“一个Excel，多个以人名命名的Sheet”
（保留样式，带进度条，原子化保存，找不到文件可弹框选择，异常写 run.log 并弹窗）
"""
import argparse
import os
import re
import sys
import tempfile
import traceback
import datetime as _dt
from collections import OrderedDict
from copy import copy
from typing import Optional, List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

DEFAULT_NAME_KEYS = ["销售员", "姓名", "员工", "人员", "负责人", "Name", "name"]

# ---------- 小工具 ----------
def exe_dir() -> str:
    # EXE/脚本所在目录
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def log_file_path() -> str:
    return os.path.join(exe_dir(), "run.log")

def log(msg: str):
    ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}\n"
    try:
        with open(log_file_path(), "a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass
    print(line, end="", flush=True)

def show_error(title: str, message: str):
    # Windows 弹窗
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(None, message, title, 0x10)
        return
    except Exception:
        pass
    # 其他平台 fallback：尝试 Tk 弹窗，否则打印
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk(); root.withdraw()
        messagebox.showerror(title, message)
        root.destroy()
    except Exception:
        print(f"{title}: {message}", file=sys.stderr)

def ask_xlsx_file() -> Optional[str]:
    # 让用户选一个 .xlsx
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw()
        path = filedialog.askopenfilename(
            title="请选择汇总Excel文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        root.destroy()
        return path if path else None
    except Exception:
        return None

def find_default_excel() -> Optional[str]:
    # 扫描 EXE 目录寻找第一个 .xlsx
    folder = exe_dir()
    for name in os.listdir(folder):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            if "按人拆分" in name:  # 避免误选输出文件
                continue
            return os.path.join(folder, name)
    return None

def detect_sheet(wb, sheet):
    if sheet is None:
        return wb.worksheets[0]
    if isinstance(sheet, str):
        if sheet.isdigit():
            idx = int(sheet)
            return wb.worksheets[idx]
        if sheet in wb.sheetnames:
            return wb[sheet]
    if isinstance(sheet, int):
        return wb.worksheets[sheet]
    return wb.worksheets[0]

def detect_name_col(header_cells: List[str], manual: Optional[str] = None) -> str:
    if manual and manual in header_cells:
        return manual
    for key in DEFAULT_NAME_KEYS:
        if key in header_cells:
            return key
    for c in header_cells:
        if any(key in c for key in DEFAULT_NAME_KEYS):
            return c
    return header_cells[0]

def base_name(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s*汇总$", "", s)
    return s.strip()

def make_sheet_title(person: str, existing: set) -> str:
    title = re.sub(r'[:\\/*?\[\]]', "_", (person or "未命名").strip() or "未命名")
    if len(title) > 31:
        title = title[:31]
    base = title
    i = 2
    while title in existing:
        suffix = f"_{i}"
        max_base_len = 31 - len(suffix)
        title = (base[:max_base_len] if max_base_len > 0 else base[:31 - len(suffix)]) + suffix
        i += 1
    return title

# ---------- 样式复制 ----------
def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)

def copy_header_and_dimensions(src_ws, dst_ws, header_row=1):
    for col_letter, dim in src_ws.column_dimensions.items():
        nd = dst_ws.column_dimensions[col_letter]
        nd.width = dim.width
        nd.hidden = dim.hidden
        nd.bestFit = dim.bestFit
    for col in range(1, src_ws.max_column + 1):
        sc = src_ws.cell(row=header_row, column=col)
        dc = dst_ws.cell(row=1, column=col)
        copy_cell(sc, dc)
    try:
        dst_ws.row_dimensions[1].height = src_ws.row_dimensions[header_row].height
    except Exception:
        pass
    dst_ws.freeze_panes = src_ws.freeze_panes

def write_row_from_src(src_ws, dst_ws, src_row_idx, dst_row_idx):
    for col in range(1, src_ws.max_column + 1):
        sc = src_ws.cell(row=src_row_idx, column=col)
        dc = dst_ws.cell(row=dst_row_idx, column=col)
        copy_cell(sc, dc)

# ---------- 原子化保存 + 自检 ----------
def safe_save_xlsx(workbook: Workbook, out_path: str):
    from openpyxl import load_workbook as _load
    folder = os.path.dirname(os.path.abspath(out_path)) or "."
    os.makedirs(folder, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix="._tmp_xlsx_", suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        workbook.save(tmp)
        _wb = _load(tmp, read_only=True, data_only=True)  # 自检
        _wb.close()
        os.replace(tmp, out_path)  # 原子替换
    finally:
        if os.path.exists(tmp):
            try: os.remove(tmp)
            except: pass

# ---------- 主逻辑 ----------
def split_to_sheets(in_path: str, sheet_sel, name_col_manual: Optional[str],
                    out_file: str, keep_empty: bool, show_progress: bool):
    log(f"输入文件：{in_path}")
    src_wb = load_workbook(in_path, data_only=True)
    src_ws = detect_sheet(src_wb, sheet_sel)
    log(f"工作表：{src_ws.title}")

    header = [str(c.value).strip() if c.value is not None else "" for c in next(src_ws.iter_rows(min_row=1, max_row=1))]
    if not header or all(not h for h in header):
        raise RuntimeError("无法读取表头（第 1 行为空）。")

    name_col = detect_name_col(header, name_col_manual)
    try:
        name_col_idx = header.index(name_col) + 1
    except ValueError:
        raise RuntimeError(f"未找到姓名列：{name_col}")
    log(f"使用姓名列：{name_col}（第 {name_col_idx} 列）")

    out_wb = Workbook()
    default_ws = out_wb.active
    out_wb.remove(default_ws)

    books: Dict[str, object] = OrderedDict()
    existing_titles = set()
    header_row_idx = 1

    total_rows = max(src_ws.max_row - 1, 0)
    pbar = tqdm(total=total_rows, desc="写入各人员Sheet", unit="行", disable=not show_progress)

    for r in range(2, src_ws.max_row + 1):
        person_raw = src_ws.cell(row=r, column=name_col_idx).value
        person = base_name(person_raw)
        if not person and not keep_empty:
            pbar.update(1); continue

        if person not in books:
            title = make_sheet_title(person, existing_titles)
            existing_titles.add(title)
            dst_ws = out_wb.create_sheet(title=title)
            copy_header_and_dimensions(src_ws, dst_ws, header_row=header_row_idx)
            books[person] = dst_ws

        dst_ws = books[person]
        dst_next_row = dst_ws.max_row + 1
        write_row_from_src(src_ws, dst_ws, r, dst_next_row)
        pbar.update(1)

    pbar.close()

    for person, ws in books.items():
        last_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    safe_save_xlsx(out_wb, out_file)
    log(f"完成！共写入 {len(books)} 个人员Sheet -> {out_file}")

def main():
    ap = argparse.ArgumentParser(description="按人名把汇总表拆成一个Excel（多Sheet，保留样式，原子化保存）")
    ap.add_argument("-i", "--input", help="输入 Excel 路径（默认自动扫描或弹框选择）")
    ap.add_argument("-s", "--sheet", help="表名或索引（0基）", default=None)
    ap.add_argument("-c", "--name-col", help="姓名列名（默认自动识别）")
    ap.add_argument("-o", "--out-file", help="输出Excel路径（默认：按人分Sheet_时间戳.xlsx）")
    ap.add_argument("--keep-empty", action="store_true", help="保留姓名为空的行（默认不保留）")
    args = ap.parse_args()

    try:
        in_path = args.input
        if not in_path or not os.path.exists(in_path):
            # 先在 EXE 目录找一个 .xlsx；没有则弹框选
            in_path = find_default_excel()
            if not in_path:
                in_path = ask_xlsx_file()
        if not in_path or not os.path.exists(in_path):
            raise FileNotFoundError("未找到输入 Excel。请把 EXE 与汇总表放同一目录，或在弹框中选择文件。")

        out_file = args.out_file or os.path.join(
            exe_dir(), f"按人分Sheet_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        # 有控制台才显示进度条
        show_progress = sys.stdout.isatty()

        split_to_sheets(in_path, args.sheet, args.name_col, out_file, args.keep_empty, show_progress)

    except Exception as e:
        tb = traceback.format_exc()
        log(f"ERROR: {e}\n{tb}")
        show_error("运行失败", f"{e}\n\n详细日志见：{log_file_path()}")

if __name__ == "__main__":
    main()
