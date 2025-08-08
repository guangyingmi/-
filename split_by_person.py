
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 汇总表一键「按人拆分 + 校验」工具（带进度条，保留样式，增强校验，报告仅 XLSX 且只标红 FAIL）
------------------------------------------------------------------------------------------------
功能：
1) 拆分：按“销售员/姓名”等列把汇总 Excel 拆成“每人一个 xlsx”，尽量保留样式：
   - 表头背景色、字体、对齐、边框、数字格式、列宽、行高、冻结窗格、自动筛选
   - 自动把“某某 汇总”行归入对应人员文件
2) 校验：立即对输出结果做一致性校验（和汇总表对比）：
   - 列头一致、行集合一致（采用“标准化+指纹+计数”比较，避免 0/0.0、日期显示差异等误报）
   - 只生成 **验证报告.xlsx**（不再生成 CSV）。其中：
       * 仅将 FAIL 行整行浅红标注；OK/WARN 不着色
       * 附“汇总”统计页
   - 可选导出差异明细：--dump-diff 时在输出目录生成
     diff_人员_src_only.csv / diff_人员_dst_only.csv
3) 进度条：拆分与校验阶段均显示进度

依赖：openpyxl>=3.1, pandas>=2.0, tqdm>=4.60
"""
import argparse
import os
import re
import sys
import hashlib
import numbers
import datetime as dt
from collections import OrderedDict, Counter
from copy import copy
import datetime as _dt
from typing import Optional, List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import pandas as pd
from tqdm import tqdm

DEFAULT_NAME_KEYS = ["销售员","姓名","员工","人员","负责人","Name","name"]


# ----------------- 日志与通用 -----------------
def log(msg: str):
    ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def find_default_excel() -> Optional[str]:
    for name in os.listdir("."):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            if "按人拆分" in name:
                continue
            return name
    return None


def detect_sheet(wb, sheet):
    if sheet is None:
        return wb.worksheets[0]
    if isinstance(sheet, str):
        if sheet.isdigit():
            idx = int(sheet)
            return wb.worksheets[idx]
        if sheet in wb.sheetnames:
            return wb[sheet]
    if isinstance(sheet, int):
        return wb.worksheets[sheet]
    return wb.worksheets[0]


def detect_name_col(header_cells: List[str], manual: Optional[str]=None) -> str:
    if manual and manual in header_cells:
        return manual
    for key in DEFAULT_NAME_KEYS:
        if key in header_cells:
            return key
    for c in header_cells:
        if any(key in c for key in DEFAULT_NAME_KEYS):
            return c
    return header_cells[0]


def base_name(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s*汇总$", "", s)  # 去掉“ 汇总”后缀
    return s.strip()


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name)


# ----------------- 样式复制 -----------------
def copy_cell(src, dst):
    """复制单元格的值与样式（字体、对齐、边框、填充、数字格式、保护）。"""
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def copy_header_and_dimensions(src_ws, dst_ws, header_row=1):
    """复制列宽、表头样式、行高、冻结窗格。"""
    for col_letter, dim in src_ws.column_dimensions.items():
        nd = dst_ws.column_dimensions[col_letter]
        nd.width = dim.width
        nd.hidden = dim.hidden
        nd.bestFit = dim.bestFit
    # 表头（源 header_row -> 目标第 1 行）
    for col in range(1, src_ws.max_column + 1):
        sc = src_ws.cell(row=header_row, column=col)
        dc = dst_ws.cell(row=1, column=col)
        copy_cell(sc, dc)
    # 行高
    try:
        dst_ws.row_dimensions[1].height = src_ws.row_dimensions[header_row].height
    except Exception:
        pass
    # 冻结窗格（如 A2）
    dst_ws.freeze_panes = src_ws.freeze_panes


def write_row_from_src(src_ws, dst_ws, src_row_idx, dst_row_idx):
    """按列逐格复制一整行（值+样式）。"""
    for col in range(1, src_ws.max_column + 1):
        sc = src_ws.cell(row=src_row_idx, column=col)
        dc = dst_ws.cell(row=dst_row_idx, column=col)
        copy_cell(sc, dc)


# ----------------- 拆分 -----------------
def split_excel(in_path: str, sheet_sel, name_col_manual: Optional[str],
                out_dir: str, keep_empty: bool, pbar: tqdm):
    wb = load_workbook(in_path, data_only=True)
    ws = detect_sheet(wb, sheet_sel)

    # 表头
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if not header or all(not h for h in header):
        raise RuntimeError("无法读取表头（第 1 行为空）。")
    name_col = detect_name_col(header, name_col_manual)
    try:
        name_col_idx = header.index(name_col) + 1
    except ValueError:
        raise RuntimeError(f"未找到姓名列：{name_col}")

    books: Dict[str, Dict] = OrderedDict()
    header_row_idx = 1

    # 拆分进度条：数据行数
    total_rows = ws.max_row - 1 if ws.max_row > 1 else 0
    pbar.reset(total=total_rows)
    pbar.set_description("拆分中")

    for r in range(2, ws.max_row + 1):
        person_raw = ws.cell(row=r, column=name_col_idx).value
        person = base_name(person_raw)
        if not person and not keep_empty:
            pbar.update(1)
            continue

        if person not in books:
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = person or "未命名"
            copy_header_and_dimensions(ws, new_ws, header_row=header_row_idx)
            books[person] = {"wb": new_wb, "ws": new_ws}

        dst_ws = books[person]["ws"]
        dst_next_row = dst_ws.max_row + 1
        write_row_from_src(ws, dst_ws, r, dst_next_row)
        pbar.update(1)

    # 保存阶段
    save_bar = tqdm(total=len(books), desc="保存文件", leave=False)
    for person, info in books.items():
        wb2, ws2 = info["wb"], info["ws"]
        last_col_letter = get_column_letter(ws2.max_column)
        ws2.auto_filter.ref = f"A1:{last_col_letter}{ws2.max_row}"
        safe = sanitize_filename(person) or "未命名"
        out_path = os.path.join(out_dir, f"{safe}.xlsx")
        wb2.save(out_path)
        save_bar.update(1)
    save_bar.close()

    return list(books.keys()), header


# ----------------- 校验（标准化 + 指纹 + 计数） -----------------
def canon(v):
    """把单元格值标准化成“稳定字符串”，避免 0 vs 0.0、日期格式差异、首尾空格等导致误报。"""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass

    # pandas Timestamp / datetime / date
    if isinstance(v, (pd.Timestamp, dt.datetime, dt.date)):
        try:
            ts = pd.to_datetime(v, errors="coerce")
        except Exception:
            ts = None
        if ts is not None and not pd.isna(ts):
            # 如果是日期或 00:00:00，统一为 YYYY-MM-DD，否则带时间
            if getattr(ts, "hour", 0) == 0 and getattr(ts, "minute", 0) == 0 and getattr(ts, "second", 0) == 0 and getattr(ts, "microsecond", 0) == 0:
                return ts.strftime("%Y-%m-%d")
            return ts.strftime("%Y-%m-%d %H:%M:%S")

    # 数字统一为通用格式（移除无意义的末尾 0）
    if isinstance(v, numbers.Number):
        if abs(v) < 1e-12:
            v = 0
        return f"{v:.15g}"

    # 其他：字符串去首尾空格
    return str(v).strip()


def row_fingerprint(series: pd.Series) -> str:
    parts = [canon(v) for v in series.tolist()]
    joined = "\x1f".join(parts)
    return hashlib.md5(joined.encode("utf-8")).hexdigest()


def build_person_map(df: pd.DataFrame, name_col: str) -> Dict[str, pd.DataFrame]:
    df = df.copy()
    df["_base_name"] = df[name_col].apply(base_name)
    return { p: sub.drop(columns=["_base_name"]) for p, sub in df.groupby("_base_name") if p }


def validate_output(sum_path: str, out_dir: str, sheet_sel, name_col_manual: Optional[str],
                    report_xlsx_path: str, vbar: tqdm, dump_diff: bool=False):
    # 读汇总
    if sheet_sel is None:
        df_sum = pd.read_excel(sum_path)
    else:
        try:
            if str(sheet_sel).isdigit():
                df_sum = pd.read_excel(sum_path, sheet_name=int(sheet_sel))
            else:
                df_sum = pd.read_excel(sum_path, sheet_name=sheet_sel)
        except Exception:
            df_sum = pd.read_excel(sum_path)

    df_sum.columns = [str(c).strip() for c in df_sum.columns]
    name_col = detect_name_col(df_sum.columns.tolist(), name_col_manual)

    src_map = build_person_map(df_sum, name_col)

    files = [f for f in os.listdir(out_dir) if f.lower().endswith(".xlsx")]
    people_in_files = [os.path.splitext(f)[0] for f in files]

    rows = []
    vbar.reset(total=len(src_map))
    vbar.set_description("校验中")

    for p, src_df in src_map.items():
        fname = None
        if p in people_in_files:
            fname = os.path.join(out_dir, f"{p}.xlsx")
        else:
            safe = sanitize_filename(p)
            if f"{safe}.xlsx" in files:
                fname = os.path.join(out_dir, f"{safe}.xlsx")
        if not fname or not os.path.exists(fname):
            rows.append({"person": p, "status": "FAIL", "detail": "未找到该人员的文件"})
            vbar.update(1)
            continue

        dst_df = pd.read_excel(fname)
        dst_df.columns = [str(c).strip() for c in dst_df.columns]

        # 列头
        if list(src_df.columns) != list(dst_df.columns):
            rows.append({"person": p, "status": "FAIL", "detail": "列头不一致"})
            vbar.update(1)
            continue

        # 指纹 + 计数（支持重复行）
        src_fp = src_df.apply(row_fingerprint, axis=1).tolist()
        dst_fp = dst_df.apply(row_fingerprint, axis=1).tolist()
        src_cnt = Counter(src_fp)
        dst_cnt = Counter(dst_fp)

        if src_cnt == dst_cnt:
            rows.append({"person": p, "status": "OK", "detail": f"行数={len(src_fp)} 完全一致"})
        else:
            missing = sum((src_cnt - dst_cnt).values())
            extra   = sum((dst_cnt - src_cnt).values())
            rows.append({"person": p, "status": "FAIL",
                         "detail": f"不一致：缺少{missing}行，多出{extra}行（src={len(src_fp)}, dst={len(dst_fp)}）"})
            if dump_diff:
                # 导出差异明细
                src_df2 = src_df.copy()
                dst_df2 = dst_df.copy()
                src_df2["_fp"] = src_df2.apply(row_fingerprint, axis=1)
                dst_df2["_fp"] = dst_df2.apply(row_fingerprint, axis=1)

                src_only_keys = list((src_cnt - dst_cnt).elements())
                dst_only_keys = list((dst_cnt - src_cnt).elements())

                src_only_rows = []
                for k in src_only_keys:
                    idx = src_df2.index[src_df2["_fp"] == k]
                    if len(idx):
                        r = src_df2.loc[idx[0]].drop(labels="_fp")
                        src_only_rows.append(r)
                        src_df2 = src_df2.drop(index=idx[0])

                dst_only_rows = []
                for k in dst_only_keys:
                    idx = dst_df2.index[dst_df2["_fp"] == k]
                    if len(idx):
                        r = dst_df2.loc[idx[0]].drop(labels="_fp")
                        dst_only_rows.append(r)
                        dst_df2 = dst_df2.drop(index=idx[0])

                if src_only_rows:
                    pd.DataFrame(src_only_rows).to_csv(
                        os.path.join(out_dir, f"diff_{sanitize_filename(p)}_src_only.csv"),
                        index=False, encoding="utf-8-sig")
                if dst_only_rows:
                    pd.DataFrame(dst_only_rows).to_csv(
                        os.path.join(out_dir, f"diff_{sanitize_filename(p)}_dst_only.csv"),
                        index=False, encoding="utf-8-sig")

        vbar.update(1)

    rep = pd.DataFrame(rows, columns=["person","status","detail"])
    # 仅写 XLSX（只标红 FAIL）
    write_colored_report(rep, report_xlsx_path)
    return rep


def write_colored_report(rep: pd.DataFrame, xlsx_path: str):
    """把报告写成 xlsx，仅对 FAIL 行染色为浅红；OK/WARN 不着色。"""
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # 明细页
        rep.to_excel(writer, index=False, sheet_name="报告")
        ws = writer.sheets["报告"]
        fill_fail = PatternFill(fill_type="solid", fgColor="FFFFC7CE")  # 浅红
        bold_font = Font(bold=True)

        # 标题行加粗
        for cell in ws[1]:
            cell.font = bold_font

        # 仅对 FAIL 行上色
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=2).value  # 第2列是 status
            if status == "FAIL":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill_fail

        # 汇总页（不着色）
        summary = rep.groupby("status")["person"].count().to_frame("count").reset_index()
        summary.to_excel(writer, index=False, sheet_name="汇总")


# ----------------- 主流程 -----------------
def main():
    ap = argparse.ArgumentParser(description="Excel 汇总表一键「按人拆分 + 校验」工具（带进度条，保留样式，增强校验，报告仅 XLSX 且只标红 FAIL）")
    ap.add_argument("-i","--input", help="输入 Excel 路径（默认自动扫描）")
    ap.add_argument("-s","--sheet", help="表名或索引（0基）", default=None)
    ap.add_argument("-c","--name-col", help="姓名列名（默认自动识别）")
    ap.add_argument("-o","--out-dir", help="输出目录（默认：按人拆分_时间戳）")
    ap.add_argument("--keep-empty", action="store_true", help="保留姓名为空的行（默认不保留）")
    ap.add_argument("--no-validate", action="store_true", help="只拆分不校验")
    ap.add_argument("--dump-diff", action="store_true", help="校验失败时导出差异明细 CSV（src_only/dst_only）")
    args = ap.parse_args()

    in_path = args.input or find_default_excel()
    if not in_path or not os.path.exists(in_path):
        log("未找到输入 Excel。请将 EXE 与汇总表放同一目录，或用 -i 指定路径。")
        sys.exit(2)

    out_dir = args.out_dir or f"按人拆分_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    os.makedirs(out_dir, exist_ok=True)

    log(f"输入文件：{in_path}")
    log(f"输出目录：{out_dir}")

    # 拆分
    with tqdm(total=0, leave=True) as pbar:
        people, _ = split_excel(in_path, args.sheet, args.name_col, out_dir, args.keep_empty, pbar)

    log(f"拆分完成，共生成 {len(people)} 个文件。")

    if args.no_validate:
        log("按参数要求跳过校验。")
        return

    # 仅生成 Excel 报告
    report_xlsx = os.path.join(out_dir, "验证报告.xlsx")
    with tqdm(total=0, leave=True) as vbar:
        rep = validate_output(in_path, out_dir, args.sheet, args.name_col, report_xlsx, vbar, dump_diff=args.dump_diff)

    summary = rep.groupby("status")["person"].count().to_dict()
    log(f"校验完成 -> {report_xlsx}")
    log(f"结果统计：{summary}")

    try:
        preview = rep.head(10).to_string(index=False)
        print("\n报告预览（前10行）：\n" + preview)
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"发生错误：{e}")
        raise
